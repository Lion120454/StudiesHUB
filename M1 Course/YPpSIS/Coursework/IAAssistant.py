import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import numpy as np
import os
import re
from datetime import datetime
import json
from enum import Enum

# Попытка импорта библиотек для Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Для экспорта в Excel установите: pip install openpyxl")


# ============================================================
# Класс для представления жизненного цикла разработки
# ============================================================

class SoftwareLifecycle(Enum):
    """Жизненные циклы разработки ПО"""
    WATERFALL = "waterfall"
    AGILE_SCRUM = "agile_scrum"
    KANBAN = "kanban"
    SPIRAL = "spiral"
    V_MODEL = "v_model"
    DEVOPS = "devops"
    
    @classmethod
    def get_display_names(cls):
        """Получение отображаемых имен"""
        return {
            cls.WATERFALL: "🏔️ Водопадная (Waterfall)",
            cls.AGILE_SCRUM: "🔄 Agile/Scrum",
            cls.KANBAN: "📋 Канбан (Kanban)",
            cls.SPIRAL: "🌀 Спиральная (Spiral)",
            cls.V_MODEL: "✅ V-образная (V-Model)",
            cls.DEVOPS: "⚙️ DevOps"
        }
    
    @classmethod
    def get_phases(cls, lifecycle):
        """Получение стандартных фаз для жизненного цикла"""
        phases = {
            cls.WATERFALL: [
                "📋 Анализ требований",
                "🎨 Проектирование",
                "💻 Реализация",
                "🧪 Тестирование",
                "🚀 Внедрение",
                "📅 Сопровождение"
            ],
            cls.AGILE_SCRUM: [
                "📊 Формирование Product Backlog",
                "🏃 Sprint Planning",
                "💻 Sprint Development",
                "🧪 Sprint Review & Testing",
                "🔄 Sprint Retrospective",
                "🚀 Релиз"
            ],
            cls.KANBAN: [
                "📝 Backlog",
                "🔍 Анализ",
                "💻 In Progress",
                "🧪 Review/Test",
                "✅ Done",
                "📊 Мониторинг метрик"
            ],
            cls.SPIRAL: [
                "🎯 Определение целей",
                "⚠️ Оценка рисков",
                "💻 Разработка и тестирование",
                "📋 Планирование следующей итерации"
            ],
            cls.V_MODEL: [
                "📋 Анализ требований",
                "🎨 Системное проектирование",
                "🏛️ Архитектурное проектирование",
                "💻 Модульная разработка",
                "🧪 Модульное тестирование",
                "🔗 Интеграционное тестирование",
                "✅ Системное тестирование",
                "🎯 Приемочное тестирование"
            ],
            cls.DEVOPS: [
                "📝 Plan",
                "💻 Code",
                "🔨 Build",
                "🧪 Test",
                "🚀 Release",
                "⚙️ Deploy",
                "📊 Operate",
                "🔄 Monitor"
            ]
        }
        return phases.get(lifecycle, phases[cls.WATERFALL])


# ============================================================
# Упрощенная нейронная сеть
# ============================================================

class SimpleNeuralNetwork:
    def __init__(self, input_size, hidden_size, output_size):
        self.input_size = input_size
        self.hidden_size = hidden_size
        self.output_size = output_size
        
        self.W1 = np.random.randn(input_size, hidden_size) * 0.01
        self.b1 = np.zeros((1, hidden_size))
        self.W2 = np.random.randn(hidden_size, output_size) * 0.01
        self.b2 = np.zeros((1, output_size))
        
    def forward(self, X):
        self.z1 = np.dot(X, self.W1) + self.b1
        self.a1 = np.tanh(self.z1)
        self.z2 = np.dot(self.a1, self.W2) + self.b2
        self.a2 = self.softmax(self.z2)
        return self.a2
    
    def softmax(self, x):
        exp_x = np.exp(x - np.max(x, axis=1, keepdims=True))
        return exp_x / np.sum(exp_x, axis=1, keepdims=True)
    
    def train_step(self, X, y, learning_rate=0.01):
        output = self.forward(X)
        m = X.shape[0]
        
        dz2 = output - y
        dW2 = np.dot(self.a1.T, dz2) / m
        db2 = np.sum(dz2, axis=0, keepdims=True) / m
        
        da1 = np.dot(dz2, self.W2.T)
        dz1 = da1 * (1 - np.power(self.a1, 2))
        dW1 = np.dot(X.T, dz1) / m
        db1 = np.sum(dz1, axis=0, keepdims=True) / m
        
        self.W2 -= learning_rate * dW2
        self.b2 -= learning_rate * db2
        self.W1 -= learning_rate * dW1
        self.b1 -= learning_rate * db1


# ============================================================
# Основной ИИ-класс
# ============================================================

class TaskDecompositionAI:
    def __init__(self):
        self.model = None
        self.word_to_idx = {}
        self.idx_to_word = {}
        self.max_seq_length = 50
        
        self.roles = [
            'Аналитик', 'Разработчик Backend', 'Разработчик Frontend',
            'Разработчик Mobile', 'Тестировщик QA', 'DevOps Инженер',
            'Project Manager', 'UX/UI Дизайнер', 'Data Engineer',
            'ML Инженер', 'Security Specialist', 'Technical Writer',
            'Product Owner', 'Scrum Master', 'Business Analyst'
        ]
        
        self.task_categories = {
            'web_development': ['сайт', 'веб', 'web', 'frontend', 'backend', 'интернет-магазин'],
            'mobile_development': ['мобильн', 'mobile', 'ios', 'android', 'react native', 'flutter'],
            'database': ['база данн', 'database', 'sql', 'postgresql', 'mysql', 'mongodb'],
            'devops': ['devops', 'ci/cd', 'деплой', 'kubernetes', 'docker', 'мониторинг'],
            'ai_ml': ['искусственн', 'нейросет', 'machine learning', 'chat bot', 'nlp', 'ai'],
            'analytics': ['аналитик', 'analytics', 'дашборд', 'dashboard', 'отчет', 'визуализац'],
            'security': ['безопасност', 'security', 'authentication', 'authorization'],
            'testing': ['тестиров', 'testing', 'qa', 'unit test', 'автотест']
        }
        
        self.current_lifecycle = SoftwareLifecycle.WATERFALL
        self.load_or_create_model()
    
    def load_or_create_model(self):
        model_file = 'simple_model.npz'
        vocab_file = 'vocab.json'
        
        if os.path.exists(model_file) and os.path.exists(vocab_file):
            try:
                data = np.load(model_file)
                self.model = SimpleNeuralNetwork(self.max_seq_length, 64, len(self.task_categories))
                self.model.W1 = data['W1']
                self.model.b1 = data['b1']
                self.model.W2 = data['W2']
                self.model.b2 = data['b2']
                
                with open(vocab_file, 'r', encoding='utf-8') as f:
                    vocab_data = json.load(f)
                    self.word_to_idx = vocab_data['word_to_idx']
                    self.idx_to_word = vocab_data['idx_to_word']
            except:
                self.initialize_model()
                self.train_on_examples()
        else:
            self.initialize_model()
            self.train_on_examples()
    
    def initialize_model(self):
        self.model = SimpleNeuralNetwork(self.max_seq_length, 64, len(self.task_categories))
        
        all_words = []
        for category, words in self.task_categories.items():
            all_words.extend(words)
        
        common_words = ['создать', 'разработать', 'настроить', 'реализовать',
                       'интегрировать', 'оптимизировать', 'протестировать']
        all_words.extend(common_words)
        
        for i, word in enumerate(set(all_words)):
            self.word_to_idx[word] = i + 1
            self.idx_to_word[i + 1] = word
    
    def text_to_vector(self, text):
        words = re.findall(r'\b[а-яa-z]+\b', text.lower())
        vector = np.zeros(self.max_seq_length)
        
        for i, word in enumerate(words[:self.max_seq_length]):
            vector[i] = self.word_to_idx.get(word, 0)
        
        return vector.reshape(1, -1)
    
    def train_on_examples(self):
        training_data = [
            ("Создать веб-сайт для интернет-магазина", "web_development"),
            ("Разработать мобильное приложение для заказа такси", "mobile_development"),
            ("Настроить базу данных для учета сотрудников", "database"),
            ("Настроить CI/CD пайплайн для автоматического деплоя", "devops"),
            ("Разработать чат-бота с ИИ для поддержки клиентов", "ai_ml"),
            ("Создать дашборд для аналитики продаж", "analytics"),
            ("Настроить систему безопасности и авторизации", "security"),
            ("Написать автотесты для модуля оплаты", "testing"),
        ]
        
        X_train = []
        y_train = []
        categories_list = list(self.task_categories.keys())
        
        for text, category in training_data:
            vector = self.text_to_vector(text)
            X_train.append(vector)
            y = np.zeros(len(categories_list))
            y[categories_list.index(category)] = 1
            y_train.append(y)
        
        X_train = np.vstack(X_train)
        y_train = np.array(y_train)
        
        for epoch in range(100):
            loss = self.model.train_step(X_train, y_train, learning_rate=0.01)
        
        np.savez('simple_model.npz',
                W1=self.model.W1, b1=self.model.b1,
                W2=self.model.W2, b2=self.model.b2)
        
        with open('vocab.json', 'w', encoding='utf-8') as f:
            json.dump({
                'word_to_idx': self.word_to_idx,
                'idx_to_word': self.idx_to_word
            }, f, ensure_ascii=False, indent=2)
    
    def set_lifecycle(self, lifecycle):
        self.current_lifecycle = lifecycle
    
    def identify_task_type(self, text):
        vector = self.text_to_vector(text)
        predictions = self.model.forward(vector)[0]
        categories_list = list(self.task_categories.keys())
        predicted_idx = np.argmax(predictions)
        
        text_lower = text.lower()
        for category, keywords in self.task_categories.items():
            for keyword in keywords:
                if keyword in text_lower:
                    return category
        return categories_list[predicted_idx] if predictions[predicted_idx] > 0.3 else 'web_development'
    
    def estimate_complexity(self, text):
        text_lower = text.lower()
        complexity_score = 0
        
        high_keywords = ['микросервис', 'распределенн', 'кластер', 'кубернет', 'высоконагруженн']
        medium_keywords = ['интеграция', 'оптимизация', 'резервный', 'автоматизация']
        
        for word in high_keywords:
            if word in text_lower:
                complexity_score += 2
        for word in medium_keywords:
            if word in text_lower:
                complexity_score += 1
        
        if complexity_score >= 3:
            return "Высокая (2-4 недели)"
        elif complexity_score >= 1:
            return "Средняя (1-2 недели)"
        else:
            return "Низкая (3-7 дней)"
    
    def assign_roles(self, text, task_type, lifecycle):
        text_lower = text.lower()
        
        role_keywords = {
            'Аналитик': ['анализ', 'требование', 'спецификаци'],
            'Разработчик Backend': ['api', 'сервер', 'бэкенд', 'база данн'],
            'Разработчик Frontend': ['интерфейс', 'frontend', 'ui'],
            'Тестировщик QA': ['тест', 'qa', 'качество', 'ошибк'],
            'DevOps Инженер': ['devops', 'деплой', 'ci/cd'],
            'Project Manager': ['менеджер', 'управление', 'координаци'],
            'Scrum Master': ['scrum', 'agile', 'команда'],
            'Product Owner': ['product', 'владелец', 'требовани']
        }
        
        scores = {role: 0 for role in self.roles}
        
        for role, keywords in role_keywords.items():
            for keyword in keywords:
                if keyword in text_lower or keyword in task_type:
                    scores[role] += 1
        
        if lifecycle in [SoftwareLifecycle.AGILE_SCRUM]:
            if 'Scrum Master' not in scores:
                scores['Scrum Master'] = 2
            if 'Product Owner' not in scores:
                scores['Product Owner'] = 2
        elif lifecycle == SoftwareLifecycle.DEVOPS:
            scores['DevOps Инженер'] += 2
        
        selected = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:6]
        return [role for role, score in selected if score > 0] or self.roles[:4]
    
    def generate_subtasks_by_lifecycle(self, text, task_type, lifecycle, roles):
        text_lower = text.lower()
        phases = SoftwareLifecycle.get_phases(lifecycle)
        
        subtasks = []
        
        if "Анализ" in str(phases[0]) or "Plan" in str(phases[0]) or "Backlog" in str(phases[0]):
            analysis_tasks = self.get_analysis_tasks(text_lower, task_type, lifecycle)
            subtasks.append({
                'phase': f"📋 {phases[0]}",
                'tasks': analysis_tasks,
                'duration': self.get_phase_duration(lifecycle, 0),
                'responsible': [r for r in roles if 'Аналитик' in r or 'Product Owner' in r][:2]
            })
        
        if len(phases) > 1:
            design_tasks = self.get_design_tasks(text_lower, task_type, lifecycle)
            subtasks.append({
                'phase': f"🎨 {phases[1]}",
                'tasks': design_tasks,
                'duration': self.get_phase_duration(lifecycle, 1),
                'responsible': [r for r in roles if 'Дизайнер' in r or 'Разработчик' in r][:2]
            })
        
        if len(phases) > 2:
            dev_tasks = self.get_development_tasks(text_lower, task_type, lifecycle)
            subtasks.append({
                'phase': f"💻 {phases[2]}",
                'tasks': dev_tasks,
                'duration': self.get_phase_duration(lifecycle, 2),
                'responsible': [r for r in roles if 'Разработчик' in r][:3]
            })
        
        if len(phases) > 3:
            test_tasks = self.get_testing_tasks(text_lower, task_type, lifecycle)
            subtasks.append({
                'phase': f"🧪 {phases[3]}",
                'tasks': test_tasks,
                'duration': self.get_phase_duration(lifecycle, 3),
                'responsible': [r for r in roles if 'Тестировщик' in r or 'QA' in r][:2]
            })
        
        if len(phases) > 4:
            deploy_tasks = self.get_deployment_tasks(text_lower, task_type, lifecycle)
            subtasks.append({
                'phase': f"🚀 {phases[4]}",
                'tasks': deploy_tasks,
                'duration': self.get_phase_duration(lifecycle, 4),
                'responsible': [r for r in roles if 'DevOps' in r or 'Разработчик' in r][:2]
            })
        
        if len(phases) > 5:
            extra_tasks = self.get_extra_phase_tasks(text_lower, lifecycle, phases[5:])
            for i, extra_phase in enumerate(phases[5:]):
                if i < len(extra_tasks):
                    subtasks.append({
                        'phase': f"📊 {extra_phase}",
                        'tasks': extra_tasks[i] if isinstance(extra_tasks[i], list) else [extra_tasks[i]],
                        'duration': self.get_phase_duration(lifecycle, 5 + i),
                        'responsible': roles[:2]
                    })
        
        return subtasks
    
    def get_analysis_tasks(self, text, task_type, lifecycle):
        tasks = [
            "Сбор и анализ требований заказчика",
            "Определение ключевых метрик успеха",
            "Оценка рисков и технических ограничений"
        ]
        
        if lifecycle == SoftwareLifecycle.AGILE_SCRUM:
            tasks.extend([
                "Формирование Product Backlog",
                "Определение критериев готовности (DoD)",
                "Оценка story points и приоритетов"
            ])
        elif lifecycle == SoftwareLifecycle.SPIRAL:
            tasks.extend([
                "Идентификация критических рисков",
                "Разработка прототипа для проверки концепции"
            ])
        elif lifecycle == SoftwareLifecycle.V_MODEL:
            tasks.extend([
                "Разработка требований к системе",
                "Создание спецификации требований к ПО",
                "Планирование приемочного тестирования"
            ])
        
        if 'api' in text:
            tasks.append("Анализ API сторонних сервисов")
        if 'безопасн' in text:
            tasks.append("Анализ требований к безопасности")
        
        return tasks
    
    def get_design_tasks(self, text, task_type, lifecycle):
        tasks = []
        
        if 'web' in task_type or 'сайт' in text:
            tasks.extend([
                "Создание прототипов интерфейса",
                "Разработка дизайн-системы",
                "Проектирование архитектуры БД"
            ])
        elif 'mobile' in task_type:
            tasks.extend([
                "Проектирование мобильного интерфейса",
                "Дизайн навигации",
                "Адаптация под разные устройства"
            ])
        else:
            tasks.extend([
                "Архитектурное проектирование",
                "Выбор технологического стека",
                "Создание ER-диаграммы"
            ])
        
        if lifecycle == SoftwareLifecycle.V_MODEL:
            tasks.extend([
                "Системное проектирование",
                "Проектирование архитектуры",
                "Детальное проектирование модулей"
            ])
        
        return tasks
    
    def get_development_tasks(self, text, task_type, lifecycle):
        tasks = [
            "Настройка окружения разработки",
            "Создание репозитория и ветвления",
            "Реализация основной бизнес-логики"
        ]
        
        if 'чат' in text or 'бот' in text:
            tasks.extend([
                "Настройка NLP модели",
                "Создание базы знаний",
                "Интеграция с мессенджерами"
            ])
        elif 'дашборд' in text:
            tasks.extend([
                "Настройка ETL пайплайна",
                "Создание визуализаций",
                "Реализация фильтров"
            ])
        elif 'магазин' in text:
            tasks.extend([
                "Реализация каталога товаров",
                "Разработка корзины",
                "Интеграция платежного шлюза"
            ])
        
        if lifecycle == SoftwareLifecycle.AGILE_SCRUM:
            tasks.append("Разбивка на sprint'ы (1-2 недели)")
            tasks.append("Ежедневные stand-up встречи")
        elif lifecycle == SoftwareLifecycle.KANBAN:
            tasks.append("Ограничение WIP (work in progress)")
            tasks.append("Визуализация потока задач")
        elif lifecycle == SoftwareLifecycle.DEVOPS:
            tasks.append("Инфраструктура как код (IaC)")
            tasks.append("Автоматизация сборки")
        
        return tasks[:8]
    
    def get_testing_tasks(self, text, task_type, lifecycle):
        tasks = [
            "Функциональное тестирование",
            "Регрессионное тестирование",
            "Нагрузочное тестирование"
        ]
        
        if lifecycle == SoftwareLifecycle.V_MODEL:
            tasks = [
                "Модульное тестирование",
                "Интеграционное тестирование",
                "Системное тестирование",
                "Приемочное тестирование"
            ]
        elif lifecycle == SoftwareLifecycle.AGILE_SCRUM:
            tasks.extend([
                "Демонстрация результатов спринта",
                "Приемочное тестирование"
            ])
        
        if 'безопасн' in text:
            tasks.append("Пентест и аудит безопасности")
        if 'api' in text:
            tasks.append("API тестирование (Postman/авто)")
        
        return tasks
    
    def get_deployment_tasks(self, text, task_type, lifecycle):
        tasks = [
            "Подготовка production окружения",
            "Настройка CI/CD пайплайна"
        ]
        
        if lifecycle == SoftwareLifecycle.DEVOPS:
            tasks.extend([
                "Автоматическое развертывание",
                "Настройка мониторинга",
                "Централизованное логирование",
                "Настройка алертов"
            ])
        else:
            tasks.extend([
                "Ручной деплой",
                "Проверка работоспособности",
                "План отката"
            ])
        
        if 'мониторинг' in text:
            tasks.append("Настройка системы мониторинга (Prometheus/Grafana)")
        
        return tasks
    
    def get_extra_phase_tasks(self, text, lifecycle, extra_phases):
        tasks = []
        
        if lifecycle == SoftwareLifecycle.DEVOPS:
            for phase in extra_phases:
                if 'Monitor' in phase:
                    tasks.append([
                        "Настройка метрик производительности",
                        "Сбор логов и ошибок",
                        "Алертинг критических событий",
                        "Дашборды мониторинга"
                    ])
                elif 'Operate' in phase:
                    tasks.append([
                        "Поддержка доступности сервисов",
                        "Резервное копирование",
                        "Масштабирование при нагрузке"
                    ])
                else:
                    tasks.append([f"Выполнение задач фазы {phase}"])
        
        elif lifecycle == SoftwareLifecycle.AGILE_SCRUM:
            tasks.append(["Следующий спринт планирование"])
            tasks.append(["Ретроспектива и улучшения"])
        
        return tasks
    
    def get_phase_duration(self, lifecycle, phase_index):
        durations = {
            SoftwareLifecycle.WATERFALL: ["1-2 дня", "2-3 дня", "4-7 дней", "2-3 дня", "1-2 дня", "1-2 дня"],
            SoftwareLifecycle.AGILE_SCRUM: ["1-2 дня", "0.5 дня", "10-14 дней", "1-2 дня", "0.5 дня", "1 день"],
            SoftwareLifecycle.KANBAN: ["вариативно", "1-3 дня", "3-5 дней", "1-2 дня", "0.5 дня", "ежедневно"],
            SoftwareLifecycle.SPIRAL: ["1-2 дня", "1 день", "7-14 дней", "1 день"],
            SoftwareLifecycle.V_MODEL: ["2-3 дня", "2-3 дня", "3-4 дня", "5-7 дней", "2-3 дня", "2-3 дня", "2-3 дня", "1-2 дня"],
            SoftwareLifecycle.DEVOPS: ["непрерывно", "непрерывно", "непрерывно", "автоматически", "непрерывно", "непрерывно", "непрерывно", "непрерывно"]
        }
        default = durations.get(lifecycle, durations[SoftwareLifecycle.WATERFALL])
        return default[phase_index] if phase_index < len(default) else "по плану"
    
    def decompose_task(self, task_description):
        task_type = self.identify_task_type(task_description)
        complexity = self.estimate_complexity(task_description)
        roles = self.assign_roles(task_description, task_type, self.current_lifecycle)
        subtasks = self.generate_subtasks_by_lifecycle(
            task_description, task_type, self.current_lifecycle, roles
        )
        
        type_names = {
            'web_development': '🌐 Веб-разработка',
            'mobile_development': '📱 Мобильная разработка',
            'database': '🗄️ Базы данных',
            'devops': '⚙️ DevOps',
            'ai_ml': '🧠 AI/ML разработка',
            'analytics': '📊 Аналитика',
            'security': '🔒 Безопасность',
            'testing': '🧪 Тестирование'
        }
        
        lifecycle_names = SoftwareLifecycle.get_display_names()
        
        return {
            'task_type': type_names.get(task_type, 'Разработка'),
            'complexity': complexity,
            'lifecycle': lifecycle_names.get(self.current_lifecycle, 'Водопадная'),
            'roles': roles,
            'subtasks': subtasks,
            'raw_task': task_description
        }


# ============================================================
# Класс для экспорта в Excel
# ============================================================

class ExcelExporter:
    """Экспорт результатов декомпозиции в Excel с форматированием"""
    
    @staticmethod
    def export_to_excel(result, filename):
        """Экспорт результата в Excel файл"""
        if not EXCEL_AVAILABLE:
            raise ImportError("Библиотека openpyxl не установлена. Установите: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Создаем листы
        summary_sheet = wb.create_sheet("Общая информация", 0)
        roles_sheet = wb.create_sheet("Роли и ответственность", 1)
        subtasks_sheet = wb.create_sheet("Декомпозиция подзадач", 2)
        timeline_sheet = wb.create_sheet("Таймлайн проекта", 3)
        
        # Заполняем все листы
        ExcelExporter._fill_summary_sheet(summary_sheet, result)
        ExcelExporter._fill_roles_sheet(roles_sheet, result)
        ExcelExporter._fill_subtasks_sheet(subtasks_sheet, result)
        ExcelExporter._fill_timeline_sheet(timeline_sheet, result)
        
        # Сохраняем файл
        wb.save(filename)
    
    @staticmethod
    def _fill_summary_sheet(sheet, result):
        """Заполнение листа с общей информацией"""
        # Стили
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0066cc', end_color='0066cc', fill_type='solid')
        title_font = Font(name='Arial', size=12, bold=True)
        normal_font = Font(name='Arial', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        sheet.merge_cells('A1:B1')
        cell = sheet['A1']
        cell.value = "📊 ИТОГОВЫЙ ОТЧЕТ ПО ДЕКОМПОЗИЦИИ ЗАДАЧИ"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.row_dimensions[1].height = 30
        
        # Дата
        sheet['A3'] = "Дата создания:"
        sheet['B3'] = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        sheet['A3'].font = title_font
        sheet['B3'].font = normal_font
        
        # Информация о задаче
        sheet['A5'] = "📝 ИСХОДНАЯ ЗАДАЧА:"
        sheet['A5'].font = title_font
        sheet.merge_cells('A5:B5')
        
        sheet.merge_cells('A6:B8')
        cell = sheet['A6']
        cell.value = result['raw_task']
        cell.font = normal_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = border
        sheet.row_dimensions[6].height = 60
        
        # Результаты анализа
        sheet['A10'] = "📌 РЕЗУЛЬТАТЫ АНАЛИЗА:"
        sheet['A10'].font = title_font
        sheet.merge_cells('A10:B10')
        
        analysis_data = [
            ("Тип задачи:", result['task_type']),
            ("Сложность:", result['complexity']),
            ("Жизненный цикл:", result['lifecycle']),
            ("Количество этапов:", str(len(result['subtasks']))),
            ("Общее количество подзадач:", str(sum(len(p['tasks']) for p in result['subtasks']))),
            ("Количество ролей:", str(len(result['roles'])))
        ]
        
        for i, (label, value) in enumerate(analysis_data):
            row = 11 + i
            sheet[f'A{row}'] = label
            sheet[f'B{row}'] = value
            sheet[f'A{row}'].font = Font(bold=True)
            sheet[f'A{row}'].border = border
            sheet[f'B{row}'].border = border
        
        # Настройка ширины колонок
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 50
    
    @staticmethod
    def _fill_roles_sheet(sheet, result):
        """Заполнение листа с ролями"""
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4caf50', end_color='4caf50', fill_type='solid')
        subheader_font = Font(name='Arial', size=12, bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        sheet.merge_cells('A1:C1')
        cell = sheet['A1']
        cell.value = "👥 РАСПРЕДЕЛЕНИЕ РОЛЕЙ И ОТВЕТСТВЕННОСТИ"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        sheet.row_dimensions[1].height = 30
        
        # Заголовки таблицы
        headers = ['№', 'Роль', 'Основные обязанности']
        role_descriptions = {
            'Аналитик': 'Сбор и анализ требований, составление ТЗ, коммуникация с заказчиком',
            'Разработчик Backend': 'Реализация серверной логики, API, работа с БД',
            'Разработчик Frontend': 'Создание пользовательского интерфейса, клиентская логика',
            'Разработчик Mobile': 'Разработка мобильных приложений под iOS/Android',
            'Тестировщик QA': 'Тестирование качества, написание автотестов, поиск багов',
            'DevOps Инженер': 'Настройка инфраструктуры, CI/CD, мониторинг',
            'Project Manager': 'Управление проектом, координация команды, отчетность',
            'UX/UI Дизайнер': 'Дизайн интерфейсов, создание прототипов',
            'Data Engineer': 'Работа с данными, ETL процессы, оптимизация запросов',
            'ML Инженер': 'Разработка моделей машинного обучения',
            'Security Specialist': 'Аудит безопасности, защита данных',
            'Technical Writer': 'Написание технической документации',
            'Product Owner': 'Управление требованиями, приоритезация бэклога',
            'Scrum Master': 'Организация Scrum-процессов, facilitation',
            'Business Analyst': 'Бизнес-анализ, перевод требований в задачи'
        }
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col)
            cell.value = header
            cell.font = subheader_font
            cell.fill = PatternFill(start_color='e0e0e0', end_color='e0e0e0', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for i, role in enumerate(result['roles'], 1):
            sheet.cell(row=2 + i, column=1, value=i).border = border
            sheet.cell(row=2 + i, column=2, value=role).border = border
            sheet.cell(row=2 + i, column=3, value=role_descriptions.get(role, 'Участие в проекте согласно компетенциям')).border = border
            sheet.cell(row=2 + i, column=1).alignment = Alignment(horizontal='center')
        
        # Настройка ширины
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 55
    
    @staticmethod
    def _fill_subtasks_sheet(sheet, result):
        """Заполнение листа с декомпозицией подзадач"""
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='ff9800', end_color='ff9800', fill_type='solid')
        phase_font = Font(name='Arial', size=12, bold=True)
        phase_fill = PatternFill(start_color='e3f2fd', end_color='e3f2fd', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        sheet.merge_cells('A1:E1')
        cell = sheet['A1']
        cell.value = "📋 ДЕТАЛЬНАЯ ДЕКОМПОЗИЦИЯ ПОДЗАДАЧ"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        sheet.row_dimensions[1].height = 30
        
        # Заголовки таблицы
        headers = ['№ этапа', 'Этап', '№ задачи', 'Подзадача', 'Срок', 'Ответственные']
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row = 3
        for phase_idx, phase_data in enumerate(result['subtasks'], 1):
            # Заголовок этапа
            sheet.merge_cells(f'A{row}:E{row}')
            cell = sheet.cell(row=row, column=1)
            cell.value = f"✨ {phase_data['phase']}"
            cell.font = phase_font
            cell.fill = phase_fill
            cell.border = border
            
            # Подзадачи этапа
            for task_idx, task in enumerate(phase_data['tasks'], 1):
                row += 1
                sheet.cell(row=row, column=1, value=phase_idx).border = border
                sheet.cell(row=row, column=1).alignment = Alignment(horizontal='center')
                sheet.cell(row=row, column=2, value=phase_data['phase']).border = border
                sheet.cell(row=row, column=3, value=task_idx).border = border
                sheet.cell(row=row, column=3).alignment = Alignment(horizontal='center')
                sheet.cell(row=row, column=4, value=task).border = border
                sheet.cell(row=row, column=5, value=phase_data['duration']).border = border
                sheet.cell(row=row, column=6, value=', '.join(phase_data['responsible'])).border = border
            
            row += 1  # Пустая строка между этапами
        
        # Настройка ширины
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 8
        sheet.column_dimensions['D'].width = 55
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 30
    
    @staticmethod
    def _fill_timeline_sheet(sheet, result):
        """Заполнение листа с таймлайном проекта"""
        header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='9c27b0', end_color='9c27b0', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        sheet.merge_cells('A1:D1')
        cell = sheet['A1']
        cell.value = "📅 ТАЙМЛАЙН ПРОЕКТА"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        sheet.row_dimensions[1].height = 30
        
        # Заголовки
        headers = ['№', 'Этап', 'Длительность', 'Количество подзадач']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=2, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='f5f5f5', end_color='f5f5f5', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for i, phase_data in enumerate(result['subtasks'], 1):
            sheet.cell(row=2 + i, column=1, value=i).border = border
            sheet.cell(row=2 + i, column=1).alignment = Alignment(horizontal='center')
            sheet.cell(row=2 + i, column=2, value=phase_data['phase'].replace('📋', '').replace('🎨', '').replace('💻', '').replace('🧪', '').strip()).border = border
            sheet.cell(row=2 + i, column=3, value=phase_data['duration']).border = border
            sheet.cell(row=2 + i, column=4, value=len(phase_data['tasks'])).border = border
            sheet.cell(row=2 + i, column=4).alignment = Alignment(horizontal='center')
        
        # Общая статистика
        total_tasks = sum(len(p['tasks']) for p in result['subtasks'])
        total_phases = len(result['subtasks'])
        
        sheet.cell(row=2 + len(result['subtasks']) + 2, column=1, value="ИТОГО:")
        sheet.cell(row=2 + len(result['subtasks']) + 2, column=1).font = Font(bold=True)
        sheet.cell(row=2 + len(result['subtasks']) + 2, column=3, value=f"{total_phases} этапов")
        sheet.cell(row=2 + len(result['subtasks']) + 3, column=3, value=f"{total_tasks} подзадач")
        
        # Настройка ширины
        sheet.column_dimensions['A'].width = 6
        sheet.column_dimensions['B'].width = 40
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 18


# ============================================================
# Графический интерфейс
# ============================================================

class DeepSeekAIAssistant:
    def __init__(self, root):
        self.root = root
        self.root.title("DeepSeek AI Assistant - Task Decomposer v3.0 (Excel Export)")
        self.root.geometry("1300x850")
        self.root.configure(bg='#0a0a0a')
        
        self.ai = TaskDecompositionAI()
        self.current_result = None
        
        self.setup_ui()
        self.apply_dark_theme()
    
    def apply_dark_theme(self):
        style = ttk.Style()
        style.theme_use('clam')
        
        style.configure('TLabel', background='#0a0a0a', foreground='#e0e0e0', font=('Segoe UI', 10))
        style.configure('TFrame', background='#0a0a0a')
        style.configure('TLabelframe', background='#0a0a0a', foreground='#e0e0e0')
        style.configure('TLabelframe.Label', background='#0a0a0a', foreground='#0066cc', font=('Segoe UI', 10, 'bold'))
    
    def setup_ui(self):
        main_container = tk.Frame(self.root, bg='#0a0a0a')
        main_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Заголовок
        header_frame = tk.Frame(main_container, bg='#0a0a0a')
        header_frame.pack(fill='x', pady=(0, 20))
        
        title_label = tk.Label(
            header_frame,
            text="🧠 DeepSeek AI Assistant - Intelligent Task Decomposer",
            font=('Segoe UI', 20, 'bold'),
            bg='#0a0a0a',
            fg='#0066cc'
        )
        title_label.pack()
        
        subtitle_label = tk.Label(
            header_frame,
            text="Автоматическая декомпозиция задач с выбором жизненного цикла разработки",
            font=('Segoe UI', 11),
            bg='#0a0a0a',
            fg='#888888'
        )
        subtitle_label.pack()
        
        # Основной контент
        content_frame = tk.Frame(main_container, bg='#0a0a0a')
        content_frame.pack(fill='both', expand=True)
        
        # Левая панель
        left_panel = tk.Frame(content_frame, bg='#0a0a0a')
        left_panel.pack(side='left', fill='both', expand=True, padx=(0, 10))
        
        # Выбор жизненного цикла
        lifecycle_frame = ttk.LabelFrame(left_panel, text="🔄 Выбор жизненного цикла разработки", padding=10)
        lifecycle_frame.pack(fill='x', pady=(0, 10))
        
        self.lifecycle_var = tk.StringVar(value="waterfall")
        
        lifecycle_options = [
            ("🏔️ Водопадная (Waterfall)", "waterfall"),
            ("🔄 Agile/Scrum", "agile_scrum"),
            ("📋 Канбан (Kanban)", "kanban"),
            ("🌀 Спиральная (Spiral)", "spiral"),
            ("✅ V-образная (V-Model)", "v_model"),
            ("⚙️ DevOps", "devops")
        ]
        
        lifecycle_inner = tk.Frame(lifecycle_frame, bg='#0a0a0a')
        lifecycle_inner.pack(fill='x')
        
        for i, (text, value) in enumerate(lifecycle_options):
            rb = tk.Radiobutton(
                lifecycle_inner,
                text=text,
                value=value,
                variable=self.lifecycle_var,
                bg='#0a0a0a',
                fg='#e0e0e0',
                selectcolor='#0a0a0a',
                activebackground='#0a0a0a',
                activeforeground='#0066cc',
                font=('Segoe UI', 10),
                command=self.on_lifecycle_change
            )
            rb.pack(side='left', padx=10)
        
        # Описание задачи
        input_label = tk.Label(
            left_panel,
            text="📝 Описание задачи",
            font=('Segoe UI', 12, 'bold'),
            bg='#0a0a0a',
            fg='#0066cc'
        )
        input_label.pack(anchor='w', pady=(10, 5))
        
        self.task_input = scrolledtext.ScrolledText(
            left_panel,
            height=12,
            font=('Consolas', 11),
            bg='#1a1a1a',
            fg='#e0e0e0',
            insertbackground='#e0e0e0',
            wrap='word',
            relief='flat'
        )
        self.task_input.pack(fill='both', expand=True)
        
        # Кнопки
        button_frame = tk.Frame(left_panel, bg='#0a0a0a')
        button_frame.pack(fill='x', pady=10)
        
        self.decompose_btn = tk.Button(
            button_frame,
            text="🚀 Декомпозировать",
            command=self.decompose_task,
            font=('Segoe UI', 11, 'bold'),
            bg='#0066cc',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=20,
            pady=8
        )
        self.decompose_btn.pack(side='left', padx=5)
        
        self.clear_btn = tk.Button(
            button_frame,
            text="🗑️ Очистить",
            command=self.clear_all,
            font=('Segoe UI', 11),
            bg='#333333',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=20,
            pady=8
        )
        self.clear_btn.pack(side='left', padx=5)
        
        # Кнопка экспорта в Excel
        self.excel_btn = tk.Button(
            button_frame,
            text="📊 Экспорт в Excel",
            command=self.export_to_excel,
            font=('Segoe UI', 11, 'bold'),
            bg='#4caf50',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=20,
            pady=8
        )
        self.excel_btn.pack(side='left', padx=5)
        
        # Примеры
        examples_label = tk.Label(
            left_panel,
            text="💡 Быстрые примеры:",
            font=('Segoe UI', 10),
            bg='#0a0a0a',
            fg='#888888'
        )
        examples_label.pack(anchor='w', pady=(10, 5))
        
        examples_frame = tk.Frame(left_panel, bg='#0a0a0a')
        examples_frame.pack(fill='x')
        
        examples = [
            ("🌐 Интернет-магазин", "Создать интернет-магазин с корзиной и оплатой"),
            ("🤖 Чат-бот", "Разработать ИИ чат-бота для техподдержки"),
            ("📊 Дашборд", "Создать дашборд аналитики продаж в реальном времени")
        ]
        
        for name, desc in examples:
            btn = tk.Button(
                examples_frame,
                text=name,
                command=lambda d=desc: self.set_example(d),
                font=('Segoe UI', 9),
                bg='#1a1a1a',
                fg='#e0e0e0',
                cursor='hand2',
                relief='flat',
                padx=10,
                pady=5
            )
            btn.pack(side='left', padx=5)
        
        # Правая панель
        right_panel = tk.Frame(content_frame, bg='#0a0a0a')
        right_panel.pack(side='right', fill='both', expand=True, padx=(10, 0))
        
        # Информационная панель
        info_frame = tk.Frame(right_panel, bg='#1a1a1a', relief='flat')
        info_frame.pack(fill='x', pady=(0, 10))
        
        info_inner = tk.Frame(info_frame, bg='#1a1a1a')
        info_inner.pack(fill='x', padx=15, pady=15)
        
        self.type_label = tk.Label(
            info_inner,
            text="📌 Тип задачи: Не определен",
            font=('Segoe UI', 11, 'bold'),
            bg='#1a1a1a',
            fg='#0066cc'
        )
        self.type_label.pack(anchor='w', pady=2)
        
        self.complexity_label = tk.Label(
            info_inner,
            text="⚡ Сложность: Не определена",
            font=('Segoe UI', 11),
            bg='#1a1a1a',
            fg='#ff9800'
        )
        self.complexity_label.pack(anchor='w', pady=2)
        
        self.lifecycle_info_label = tk.Label(
            info_inner,
            text="🔄 Жизненный цикл: Не выбран",
            font=('Segoe UI', 11),
            bg='#1a1a1a',
            fg='#4caf50'
        )
        self.lifecycle_info_label.pack(anchor='w', pady=2)
        
        # Роли
        roles_frame = tk.Frame(right_panel, bg='#1a1a1a', relief='flat')
        roles_frame.pack(fill='x', pady=(0, 10))
        
        roles_title = tk.Label(
            roles_frame,
            text="👥 Назначенные роли",
            font=('Segoe UI', 11, 'bold'),
            bg='#1a1a1a',
            fg='#0066cc'
        )
        roles_title.pack(anchor='w', padx=15, pady=(15, 10))
        
        self.roles_text = tk.Text(
            roles_frame,
            height=6,
            font=('Segoe UI', 10),
            bg='#1a1a1a',
            fg='#e0e0e0',
            wrap='word',
            relief='flat'
        )
        self.roles_text.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        # Подзадачи
        subtasks_frame = tk.Frame(right_panel, bg='#1a1a1a', relief='flat')
        subtasks_frame.pack(fill='both', expand=True)
        
        subtasks_title = tk.Label(
            subtasks_frame,
            text="📋 Детальная декомпозиция подзадач",
            font=('Segoe UI', 11, 'bold'),
            bg='#1a1a1a',
            fg='#0066cc'
        )
        subtasks_title.pack(anchor='w', padx=15, pady=(15, 10))
        
        self.subtasks_text = scrolledtext.ScrolledText(
            subtasks_frame,
            font=('Consolas', 10),
            bg='#1a1a1a',
            fg='#e0e0e0',
            wrap='word',
            relief='flat'
        )
        self.subtasks_text.pack(fill='both', expand=True, padx=15, pady=(0, 15))
        
        # Статус бар
        status_frame = tk.Frame(main_container, bg='#0a0a0a')
        status_frame.pack(fill='x', pady=(20, 0))
        
        self.status_bar = tk.Label(
            status_frame,
            text="✅ Готов к работе | Выберите жизненный цикл и введите задачу",
            font=('Segoe UI', 9),
            bg='#0a0a0a',
            fg='#666666'
        )
        self.status_bar.pack(fill='x')
    
    def on_lifecycle_change(self):
        value = self.lifecycle_var.get()
        lifecycle_map = {
            "waterfall": SoftwareLifecycle.WATERFALL,
            "agile_scrum": SoftwareLifecycle.AGILE_SCRUM,
            "kanban": SoftwareLifecycle.KANBAN,
            "spiral": SoftwareLifecycle.SPIRAL,
            "v_model": SoftwareLifecycle.V_MODEL,
            "devops": SoftwareLifecycle.DEVOPS
        }
        self.ai.set_lifecycle(lifecycle_map.get(value, SoftwareLifecycle.WATERFALL))
        
        lifecycle_names = {
            "waterfall": "Водопадная",
            "agile_scrum": "Agile/Scrum",
            "kanban": "Канбан",
            "spiral": "Спиральная",
            "v_model": "V-модель",
            "devops": "DevOps"
        }
        self.status_bar.config(text=f"✅ Выбран жизненный цикл: {lifecycle_names.get(value, 'Водопадная')}")
    
    def set_example(self, example):
        self.task_input.delete("1.0", tk.END)
        self.task_input.insert("1.0", example)
        self.status_bar.config(text="💡 Пример задачи загружен")
    
    def decompose_task(self):
        task = self.task_input.get("1.0", tk.END).strip()
        
        if not task:
            messagebox.showwarning("Внимание", "Пожалуйста, введите описание задачи")
            return
        
        self.decompose_btn.config(state='disabled', text='🔄 Анализирую...')
        self.status_bar.config(text="🧠 DeepSeek AI анализирует задачу с учетом выбранного жизненного цикла...")
        
        thread = threading.Thread(target=self.process_task, args=(task,))
        thread.daemon = True
        thread.start()
    
    def process_task(self, task):
        try:
            result = self.ai.decompose_task(task)
            self.current_result = result
            self.root.after(0, self.update_results, result)
        except Exception as e:
            self.root.after(0, self.show_error, str(e))
    
    def update_results(self, result):
        self.type_label.config(text=f"📌 Тип задачи: {result['task_type']}")
        self.complexity_label.config(text=f"⚡ Сложность: {result['complexity']}")
        self.lifecycle_info_label.config(text=f"🔄 Жизненный цикл: {result['lifecycle']}")
        
        self.roles_text.delete("1.0", tk.END)
        icons = {
            'Аналитик': '📊', 'Разработчик Backend': '⚙️', 'Разработчик Frontend': '🎨',
            'Разработчик Mobile': '📱', 'Тестировщик QA': '🐛', 'DevOps Инженер': '🔧',
            'Project Manager': '📋', 'UX/UI Дизайнер': '🎨', 'Data Engineer': '🗄️',
            'ML Инженер': '🤖', 'Security Specialist': '🔒', 'Technical Writer': '📝',
            'Product Owner': '👑', 'Scrum Master': '🔄', 'Business Analyst': '💼'
        }
        
        for role in result['roles']:
            icon = icons.get(role, '👤')
            self.roles_text.insert(tk.END, f"{icon} {role}\n\n")
        
        self.subtasks_text.delete("1.0", tk.END)
        
        for phase_data in result['subtasks']:
            self.subtasks_text.insert(tk.END, f"\n{phase_data['phase']}\n", 'phase')
            self.subtasks_text.insert(tk.END, "─" * 60 + "\n", 'separator')
            
            for i, task_item in enumerate(phase_data['tasks'], 1):
                self.subtasks_text.insert(tk.END, f"  {i}. {task_item}\n", 'task')
            
            self.subtasks_text.insert(tk.END, f"\n  ⏱️ Срок: {phase_data['duration']}\n", 'duration')
            self.subtasks_text.insert(tk.END, f"  👥 Ответственные: {', '.join(phase_data['responsible'])}\n\n", 'responsible')
        
        self.subtasks_text.tag_config('phase', foreground='#0066cc', font=('Segoe UI', 11, 'bold'))
        self.subtasks_text.tag_config('separator', foreground='#333333')
        self.subtasks_text.tag_config('task', foreground='#e0e0e0', font=('Segoe UI', 10))
        self.subtasks_text.tag_config('duration', foreground='#ff9800', font=('Segoe UI', 10, 'italic'))
        self.subtasks_text.tag_config('responsible', foreground='#4caf50', font=('Segoe UI', 10, 'italic'))
        
        self.status_bar.config(text="✅ Анализ завершен! Результаты готовы")
        self.decompose_btn.config(state='normal', text='🚀 Декомпозировать')
    
    def export_to_excel(self):
        """Экспорт результатов в Excel"""
        if not self.current_result:
            messagebox.showwarning("Внимание", "Нет результатов для экспорта. Сначала выполните декомпозицию задачи.")
            return
        
        if not EXCEL_AVAILABLE:
            result = messagebox.askyesno(
                "Библиотека не установлена",
                "Для экспорта в Excel требуется установить библиотеку openpyxl.\n\n"
                "Установить сейчас? (потребуется интернет)"
            )
            if result:
                try:
                    import subprocess
                    import sys
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
                    messagebox.showinfo("Успех", "Библиотека openpyxl успешно установлена!\nПерезапустите приложение.")
                except Exception as e:
                    messagebox.showerror("Ошибка", f"Не удалось установить библиотеку:\n{e}")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("All files", "*.*")
            ],
            title="Сохранить отчет в Excel"
        )
        
        if filename:
            try:
                self.status_bar.config(text="📊 Экспорт в Excel...")
                ExcelExporter.export_to_excel(self.current_result, filename)
                self.status_bar.config(text=f"✅ Экспорт завершен! Файл сохранен: {os.path.basename(filename)}")
                messagebox.showinfo("Успех", f"Отчет успешно экспортирован в Excel:\n{filename}")
            except Exception as e:
                self.status_bar.config(text="❌ Ошибка при экспорте в Excel")
                messagebox.showerror("Ошибка", f"Не удалось экспортировать в Excel:\n{e}")
    
    def show_error(self, error):
        messagebox.showerror("Ошибка", f"Произошла ошибка:\n{error}")
        self.status_bar.config(text="❌ Ошибка при анализе задачи")
        self.decompose_btn.config(state='normal', text='🚀 Декомпозировать')
    
    
    def clear_all(self):
        self.task_input.delete("1.0", tk.END)
        self.roles_text.delete("1.0", tk.END)
        self.subtasks_text.delete("1.0", tk.END)
        self.type_label.config(text="📌 Тип задачи: Не определен")
        self.complexity_label.config(text="⚡ Сложность: Не определена")
        self.lifecycle_info_label.config(text="🔄 Жизненный цикл: Не выбран")
        self.current_result = None
        self.status_bar.config(text="✨ Все очищено | Готов к новой задаче")


# ============================================================
# Запуск приложения
# ============================================================

if __name__ == "__main__":
    root = tk.Tk()
    root.minsize(1200, 800)
    
    root.update_idletasks()
    width = 1300
    height = 850
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f'{width}x{height}+{x}+{y}')
    
    app = DeepSeekAIAssistant(root)
    root.mainloop() 